import csv
import json
from pathlib import Path

from django.apps import apps
from django.core.management.base import BaseCommand, CommandError
from django.core.serializers import serialize
from django.db.utils import OperationalError, ProgrammingError
from django.db.models import FileField
from django.utils import timezone

try:
    from openpyxl import Workbook
except ImportError:  # pragma: no cover
    Workbook = None


DEFAULT_APP_LABELS = [
    'users',
    'organizations',
    'employees',
    'crm',
    'education',
    'erp_services',
    'finance',
    'erp_documents',
    'attendance',
    'projects_v2',
    'knowledge',
    'customfields',
    'erp_notifications',
    'catalog',
    'clients',
    'services',
    'analytics',
    'leads',
    'tasks',
    'timetracking',
    'documents',
    'notifications',
    'reports',
    'support',
]

EXCLUDED_MODELS = {
    ('contenttypes', 'ContentType'),
    ('sessions', 'Session'),
    ('admin', 'LogEntry'),
}


def safe_sheet_name(label):
    cleaned = ''.join(ch if ch not in r'[]:*?/\\' else '_' for ch in label)
    return cleaned[:31] or 'sheet'


def serialize_value(value):
    if value is None:
        return ''
    if hasattr(value, 'isoformat'):
        return value.isoformat()
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def model_filename(model):
    return f'{model._meta.app_label}.{model._meta.model_name}'


class Command(BaseCommand):
    help = 'Export ManagerSL migration data to XLSX, CSV and JSON files.'

    def add_arguments(self, parser):
        parser.add_argument('--output', required=True, help='Output directory for exported files.')
        parser.add_argument(
            '--apps',
            default=','.join(DEFAULT_APP_LABELS),
            help='Comma-separated app labels to export.',
        )
        parser.add_argument(
            '--skip-empty',
            action='store_true',
            help='Do not create CSV/JSON files for empty models.',
        )

    def handle(self, *args, **options):
        if Workbook is None:
            raise CommandError('openpyxl is not installed. Install requirements.txt first.')

        output_dir = Path(options['output']).resolve()
        csv_dir = output_dir / 'csv'
        json_dir = output_dir / 'json'
        xlsx_dir = output_dir / 'xlsx'
        for directory in (csv_dir, json_dir, xlsx_dir):
            directory.mkdir(parents=True, exist_ok=True)

        requested_labels = {item.strip() for item in options['apps'].split(',') if item.strip()}
        models_by_app = {}
        for model in apps.get_models():
            label = model._meta.app_label
            model_key = (label, model.__name__)
            if label not in requested_labels or model_key in EXCLUDED_MODELS:
                continue
            models_by_app.setdefault(label, []).append(model)

        summary = {
            'generated_at': timezone.now().isoformat(),
            'apps': {},
            'total_models': 0,
            'total_rows': 0,
        }

        for app_label, model_list in sorted(models_by_app.items()):
            workbook = Workbook()
            default_sheet = workbook.active
            workbook.remove(default_sheet)
            app_summary = {}

            for model in sorted(model_list, key=lambda item: item._meta.model_name):
                qs = model._default_manager.all().order_by(model._meta.pk.name)
                try:
                    count = qs.count()
                except (OperationalError, ProgrammingError) as exc:
                    app_summary[model._meta.model_name] = {
                        'verbose_name_plural': str(model._meta.verbose_name_plural),
                        'rows': 0,
                        'skipped': True,
                        'error': str(exc),
                    }
                    self.stdout.write(self.style.WARNING(
                        f'Skipped {model._meta.label}: {exc}'
                    ))
                    continue
                if options['skip_empty'] and count == 0:
                    continue

                fields = list(model._meta.concrete_fields)
                many_to_many = list(model._meta.many_to_many)
                headers = []
                for field in fields:
                    headers.append(field.name if not field.is_relation else f'{field.name}_id')
                for field in many_to_many:
                    headers.append(f'{field.name}_ids')

                rows = []
                try:
                    for obj in qs.iterator(chunk_size=500):
                        row = []
                        for field in fields:
                            if field.is_relation:
                                value = getattr(obj, field.attname, None)
                            else:
                                value = getattr(obj, field.name, None)
                                if isinstance(field, FileField):
                                    value = getattr(value, 'name', '') if value else ''
                            row.append(serialize_value(value))
                        for field in many_to_many:
                            try:
                                value = list(getattr(obj, field.name).values_list('pk', flat=True))
                            except Exception:
                                value = []
                            row.append(json.dumps(value, ensure_ascii=False))
                        rows.append(row)
                except (OperationalError, ProgrammingError) as exc:
                    app_summary[model._meta.model_name] = {
                        'verbose_name_plural': str(model._meta.verbose_name_plural),
                        'rows': 0,
                        'skipped': True,
                        'error': str(exc),
                    }
                    self.stdout.write(self.style.WARNING(
                        f'Skipped {model._meta.label}: {exc}'
                    ))
                    continue

                basename = model_filename(model)
                csv_path = csv_dir / f'{basename}.csv'
                with csv_path.open('w', encoding='utf-8-sig', newline='') as csv_file:
                    writer = csv.writer(csv_file)
                    writer.writerow(headers)
                    writer.writerows(rows)

                json_path = json_dir / f'{basename}.json'
                try:
                    json_path.write_text(
                        serialize('json', qs, use_natural_foreign_keys=False, use_natural_primary_keys=False),
                        encoding='utf-8',
                    )
                except (OperationalError, ProgrammingError) as exc:
                    app_summary[model._meta.model_name] = {
                        'verbose_name_plural': str(model._meta.verbose_name_plural),
                        'rows': 0,
                        'skipped': True,
                        'error': str(exc),
                    }
                    self.stdout.write(self.style.WARNING(
                        f'Skipped JSON for {model._meta.label}: {exc}'
                    ))
                    continue

                sheet = workbook.create_sheet(safe_sheet_name(model._meta.model_name))
                sheet.append(headers)
                for row in rows:
                    sheet.append(row)

                app_summary[model._meta.model_name] = {
                    'verbose_name_plural': str(model._meta.verbose_name_plural),
                    'rows': count,
                    'csv': str(csv_path.relative_to(output_dir)).replace('\\', '/'),
                    'json': str(json_path.relative_to(output_dir)).replace('\\', '/'),
                }
                summary['total_models'] += 1
                summary['total_rows'] += count

            if workbook.sheetnames:
                xlsx_path = xlsx_dir / f'{app_label}.xlsx'
                workbook.save(xlsx_path)
            summary['apps'][app_label] = app_summary

        (output_dir / 'summary.json').write_text(
            json.dumps(summary, ensure_ascii=False, indent=2),
            encoding='utf-8',
        )

        self.stdout.write(self.style.SUCCESS(
            f'Exported {summary["total_rows"]} rows from {summary["total_models"]} models to {output_dir}'
        ))
